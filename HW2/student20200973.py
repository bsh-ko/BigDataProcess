#!/usr/bin/python3
import openpyxl
import math

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

len=0
row_id = 1;
for row in ws:
	if row_id != 1: 
		sum_v = ws.cell(row = row_id, column = 3).value *0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		len += 1
	row_id += 1

total = []
total.sort(key = lambda x:x[1], reverse=True)

row_id = 1;
for row in ws:
	if row_id != 1: 
		ws.cell(row = row_id, column = 8).value = 'C0'
	row_id += 1

Bcnt = 0
row_id = 1;
while Bcnt < int(len*0.7):
	ws.cell(row = total[Bcnt][0], column = 8).value ='B0'
	Bcnt += 1
	row_id += 1

Acnt = 0
row_id = 1;
while Acnt < int(len*0.3):
	ws.cell(row = total[Acnt][0], column = 8).value ='A0'
	Acnt += 1
	row_id += 1

a = 0
row_id = 1;
while a < int(Acnt*0.5):
	ws.cell(row = total[a][0], column = 8).value ='A+'
	a += 1
	row_id += 1

b = 0
row_id = 1;
while b < int(Bcnt*0.5):
	ws.cell(row = total[b][0], column = 8).value ='B+'
	b += 1
	row_id += 1

Ccnt = len - Bcnt
c = 0
row_id = 1;
while c< math.trunc(Ccnt*0.5):
	ws.cell(row = total[c][0], column = 8).value ='C+'
	c += 1
	row_id += 1

wb.save("student.xlsx")
